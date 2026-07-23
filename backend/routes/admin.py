from datetime import datetime, timezone
import csv
import io
import json
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

from config import UPLOAD_MAX_BYTES
from database import (
    assessment_configs_collection,
    assessments_collection,
    credentials_collection,
    question_banks_collection,
    reports_collection,
    schools_collection,
    students_collection,
    users_collection,
)
from services.auth_service import require_admin
from services.csv_parser import create_student_accounts, parse_tabular_file, validate_csv_rows
from services.question_seed import LIKERT_OPTIONS, SECTIONS

router = APIRouter(prefix="/api/admin", tags=["admin"])


def scoped_school_filter(user, school_id=None):
    admin_school_id = user.get('school_id')
    if admin_school_id:
        if school_id and school_id != admin_school_id:
            raise HTTPException(status_code=403, detail="Access denied for this school")
        return {"school_id": admin_school_id}
    return {"school_id": school_id} if school_id else {}


def serialize_doc(doc):
    if not doc:
        return doc
    clean = dict(doc)
    clean['_id'] = str(clean.get('_id'))
    return clean


def parse_bool(value):
    if isinstance(value, bool):
        return value
    return str(value or '').strip().lower() in {'true', '1', 'yes', 'y'}


def clean_question(row):
    normalized = {str(k).strip().lower().replace(' ', '_'): v for k, v in row.items()}
    section = str(normalized.get('section', '')).strip()
    dimension = str(normalized.get('dimension', '')).strip()
    text = str(normalized.get('text', normalized.get('question', ''))).strip()
    question_type = str(normalized.get('question_type', 'likert_5')).strip() or 'likert_5'

    if section not in SECTIONS:
        raise ValueError(f"Invalid section '{section}'")
    if not dimension or not text:
        raise ValueError("Question must include section, dimension, and text")

    options_raw = normalized.get('options', '')
    if question_type == 'likert_5':
        options = LIKERT_OPTIONS
    elif isinstance(options_raw, list):
        options = options_raw
    else:
        try:
            parsed = json.loads(str(options_raw)) if str(options_raw).strip() else []
            options = parsed if isinstance(parsed, list) else []
        except json.JSONDecodeError:
            pieces = [p.strip() for p in str(options_raw).split('|') if p.strip()]
            options = [{"id": chr(97 + i), "text": piece, "value": 0} for i, piece in enumerate(pieces)]

    if question_type == 'multiple_choice' and not options:
        raise ValueError("Multiple-choice questions need options")

    assessment_id_val = str(normalized.get('assessment_id', '')).strip() or None
    return {
        "_id": str(normalized.get('_id') or normalized.get('id') or f"q_{uuid.uuid4().hex[:10]}").strip(),
        "section": section,
        "dimension": dimension,
        "text": text,
        "question_type": question_type,
        "options": options,
        "correct_answer": str(normalized.get('correct_answer', '')).strip() or None,
        "reverse_scored": parse_bool(normalized.get('reverse_scored', False)),
        "weight": float(normalized.get('weight') or 1.0),
        "assessment_id": assessment_id_val,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }


async def read_upload(file: UploadFile):
    content = await file.read()
    if len(content) > UPLOAD_MAX_BYTES:
        raise HTTPException(status_code=413, detail="File is too large")
    return content


def _extract_school_name(file_content: bytes, filename: str) -> str:
    """Try to extract school name from the first line of the CSV/Excel file."""
    try:
        if filename.lower().endswith(('.xlsx', '.xls')):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            sheet = wb.active
            first_row = next(sheet.iter_rows(max_row=1, values_only=True), None)
            if first_row and first_row[0]:
                name = str(first_row[0]).strip().strip('"').strip()
                if len(name) > 3 and len(name) < 100 and not name[0].isdigit():
                    return name
        else:
            text = file_content.decode('utf-8-sig')
            first_line = text.split('\n', 1)[0].strip().strip('"').strip()
            if first_line and len(first_line) > 3 and len(first_line) < 100 and not first_line[0].isdigit():
                return first_line
    except Exception:
        pass
    return ''


@router.post("/upload-students")
async def upload_students(file: UploadFile = File(...), user=Depends(require_admin)):
    if not file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are accepted")

    file_content = await read_upload(file)
    rows = parse_tabular_file(file_content, file.filename)
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or has invalid format. Expected columns: Student Name, Class (e.g. V-A, XII-B)")

    valid_rows, errors = validate_csv_rows(rows)
    if not valid_rows:
        raise HTTPException(status_code=400, detail={"message": "No valid rows found", "errors": errors})

    school_id = user.get('school_id') or f"sch_{uuid.uuid4().hex[:8]}"
    existing_school = await schools_collection.find_one({"_id": school_id})
    if not existing_school:
        school_name = _extract_school_name(file_content, file.filename) or user.get('name', 'VidyaLoop School')
        await schools_collection.insert_one({
            "_id": school_id,
            "name": school_name,
            "contact_email": user.get('email', ''),
            "status": "active",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        if not user.get('school_id'):
            await users_collection.update_one({"_id": user['_id']}, {"$set": {"school_id": school_id}})

    result = await create_student_accounts(valid_rows, school_id, user['_id'])
    credentials = result['credentials']

    return {
        "message": f"Successfully created {len(credentials)} student accounts with login credentials",
        "total_processed": len(rows),
        "successful": len(credentials),
        "skipped": len(rows) - len(valid_rows),
        "errors": errors,
        "school_id": school_id,
        "batch_id": result['batch_id'],
        "credentials": credentials,
    }


@router.post("/preview-csv")
async def preview_csv(file: UploadFile = File(...), user=Depends(require_admin)):
    file_content = await read_upload(file)
    rows = parse_tabular_file(file_content, file.filename)
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or has invalid format")
    valid_rows, errors = validate_csv_rows(rows)
    school_name = _extract_school_name(file_content, file.filename)
    return {
        "total_rows": len(rows),
        "valid_rows": len(valid_rows),
        "errors": errors,
        "school_name": school_name,
        "preview": valid_rows[:20],
    }


@router.get("/schools")
async def list_schools(search: str = "", user=Depends(require_admin)):
    query = {}
    if user.get('school_id'):
        query['_id'] = user['school_id']
    if search:
        query['name'] = {"$regex": search, "$options": "i"}

    schools = await schools_collection.find(query).sort("name", 1).to_list(100)
    results = []
    for school in schools:
        sid = school['_id']
        results.append({
            **serialize_doc(school),
            "student_count": await students_collection.count_documents({"school_id": sid}),
            "completed_assessments": await assessments_collection.count_documents({"school_id": sid, "status": "completed"}),
        })
    return {"schools": results}


@router.get("/students")
async def list_students(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    search: str = "",
    class_filter: int | None = None,
    school_id: str | None = None,
    user=Depends(require_admin),
):
    query = scoped_school_filter(user, school_id)
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"username": {"$regex": search, "$options": "i"}},
            {"roll_number": {"$regex": search, "$options": "i"}},
        ]
    if class_filter:
        query["class_level"] = class_filter

    total = await students_collection.count_documents(query)
    skip = (page - 1) * limit
    students = await students_collection.find(query).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)

    for student in students:
        student['_id'] = str(student['_id'])
        user_doc = await users_collection.find_one({"_id": student.get('user_id')})
        student['username'] = student.get('username') or (user_doc.get('username') if user_doc else '')
        student['is_active'] = user_doc.get('is_active', True) if user_doc else False
        student['assessment_count'] = await assessments_collection.count_documents({"student_id": student['_id']})
        latest = await assessments_collection.find_one({"student_id": student['_id']}, sort=[("started_at", -1)])
        student['latest_assessment_status'] = latest.get('status') if latest else None

    return {"students": students, "total": total, "page": page, "limit": limit, "total_pages": (total + limit - 1) // limit}


@router.get("/students/{student_id}")
async def get_student(student_id: str, user=Depends(require_admin)):
    student = await students_collection.find_one({"_id": student_id})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    scoped_school_filter(user, student.get('school_id'))

    assessments = await assessments_collection.find({"student_id": student_id}).sort("started_at", -1).to_list(100)
    reports = await reports_collection.find({"student_id": student_id}).sort("generated_at", -1).to_list(100)
    return {"student": serialize_doc(student), "assessments": [serialize_doc(a) for a in assessments], "reports": [serialize_doc(r) for r in reports]}


@router.get("/credentials")
async def list_credentials(user=Depends(require_admin)):
    query = scoped_school_filter(user)
    batches = await credentials_collection.find(query).sort("created_at", -1).to_list(50)
    return {"batches": [{**serialize_doc(b), "count": len(b.get('credentials', []))} for b in batches]}


@router.get("/credentials/{batch_id}/download")
async def download_credentials(batch_id: str, user=Depends(require_admin)):
    batch = await credentials_collection.find_one({"_id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Credential batch not found")
    scoped_school_filter(user, batch.get('school_id'))

    output = io.StringIO()
    fieldnames = ["student_id", "name", "class_level", "section", "username", "password"]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for credential in batch.get('credentials', []):
        row_dict = {key: credential.get(key, '') for key in writer.fieldnames}
        if not row_dict.get('username'):
            student_id = credential.get('student_id')
            user_id = credential.get('user_id')
            student_doc = await students_collection.find_one({"_id": student_id}) if student_id else None
            user_doc = await users_collection.find_one({"_id": user_id}) if user_id else None
            found_username = (student_doc.get('username') if student_doc else None) or (user_doc.get('username') if user_doc else None)
            if not found_username:
                clean_n = str(credential.get('name', 'student')).lower().replace(' ', '').replace('.', '')
                clean_n = ''.join(c for c in clean_n if c.isalnum()) or 'student'
                cls = credential.get('class_level', '')
                found_username = f"{clean_n}{cls}"
            row_dict['username'] = found_username

        if not row_dict.get('password'):
            row_dict['password'] = "demo1234"
        writer.writerow(row_dict)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=vidyaloop_credentials_{batch_id}.csv"},
    )


@router.get("/credentials/{batch_id}/preview")
async def preview_credentials(batch_id: str, user=Depends(require_admin)):
    """Preview the first 20 credentials in a batch without downloading."""
    batch = await credentials_collection.find_one({"_id": batch_id})
    if not batch:
        raise HTTPException(status_code=404, detail="Credential batch not found")
    scoped_school_filter(user, batch.get('school_id'))

    all_creds = batch.get('credentials', [])
    preview = []
    for c in all_creds[:20]:
        username = c.get('username', '')
        if not username:
            student_doc = await students_collection.find_one({"_id": c.get('student_id')}) if c.get('student_id') else None
            user_doc = await users_collection.find_one({"_id": c.get('user_id')}) if c.get('user_id') else None
            username = (student_doc.get('username') if student_doc else None) or (user_doc.get('username') if user_doc else '') or ''
        preview.append({
            "name": c.get('name', ''),
            "class_level": c.get('class_level', ''),
            "section": c.get('section', ''),
            "username": username,
            "password": c.get('password', ''),
        })

    return {
        "batch_id": batch_id,
        "total": len(all_creds),
        "preview": preview,
        "created_at": batch.get('created_at', ''),
    }


@router.get("/assessments")
async def list_assessments(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    status: str = "",
    class_filter: int | None = None,
    school_id: str | None = None,
    user=Depends(require_admin),
):
    query = scoped_school_filter(user, school_id)
    if status:
        query['status'] = status

    student_ids = None
    if class_filter:
        student_query = scoped_school_filter(user, school_id)
        student_query['class_level'] = class_filter
        student_ids = [s['_id'] for s in await students_collection.find(student_query, {"_id": 1}).to_list(10000)]
        query['student_id'] = {"$in": student_ids}

    total = await assessments_collection.count_documents(query)
    assessments = await assessments_collection.find(query).sort("started_at", -1).skip((page - 1) * limit).limit(limit).to_list(limit)
    results = []
    for assessment in assessments:
        student = await students_collection.find_one({"_id": assessment.get('student_id')})
        school = await schools_collection.find_one({"_id": assessment.get('school_id')}) if assessment.get('school_id') else None
        report = await reports_collection.find_one({"assessment_id": assessment.get('_id')})
        results.append({
            **serialize_doc(assessment),
            "student": serialize_doc(student) if student else None,
            "school": serialize_doc(school) if school else None,
            "report": serialize_doc(report) if report else None,
            "report_id": report.get('_id') if report else None,
        })
    return {"assessments": results, "total": total, "page": page, "limit": limit, "total_pages": (total + limit - 1) // limit}


@router.get("/assessments/{assessment_id}/takers")
async def get_assessment_taker(assessment_id: str, user=Depends(require_admin)):
    assessment = await assessments_collection.find_one({"_id": assessment_id})
    if not assessment:
        raise HTTPException(status_code=404, detail="Assessment not found")
    scoped_school_filter(user, assessment.get('school_id'))
    student = await students_collection.find_one({"_id": assessment.get('student_id')})
    report = await reports_collection.find_one({"assessment_id": assessment_id})
    return {"assessment": serialize_doc(assessment), "student": serialize_doc(student) if student else None, "report": serialize_doc(report) if report else None}


@router.get("/question-bank")
async def list_question_bank(section: str = "", assessment_id: str = "", user=Depends(require_admin)):
    query = {}
    if section:
        query["section"] = section
    if assessment_id:
        query["assessment_id"] = assessment_id
    questions = await question_banks_collection.find(query).sort([("section", 1), ("dimension", 1)]).to_list(500)
    return {"sections": SECTIONS, "questions": [serialize_doc(q) for q in questions]}


@router.post("/question-bank/upload")
async def upload_question_bank(file: UploadFile = File(...), assessment_id: str = "", user=Depends(require_admin)):
    content = await read_upload(file)
    if file.filename.lower().endswith('.json'):
        data = json.loads(content.decode('utf-8-sig'))
        rows = data.get('questions', data) if isinstance(data, dict) else data
    elif file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
        rows = parse_tabular_file(content, file.filename)
    else:
        raise HTTPException(status_code=400, detail="Upload JSON, CSV, or Excel question files")

    if not isinstance(rows, list) or not rows:
        raise HTTPException(status_code=400, detail="No question rows found")

    upserted = 0
    errors = []
    for idx, row in enumerate(rows):
        try:
            question = clean_question(row)
            if assessment_id:
                question['assessment_id'] = assessment_id
            await question_banks_collection.update_one({"_id": question['_id']}, {"$set": question}, upsert=True)
            upserted += 1
        except Exception as exc:
            errors.append(f"Row {idx + 2}: {exc}")

    if assessment_id:
        count = await question_banks_collection.count_documents({"assessment_id": assessment_id})
        await assessment_configs_collection.update_one({"_id": assessment_id}, {"$set": {"question_count": count, "updated_at": datetime.now(timezone.utc).isoformat()}})

    if upserted == 0:
        raise HTTPException(status_code=400, detail={"message": "No valid questions found", "errors": errors})
    return {"message": f"Saved {upserted} questions", "saved": upserted, "errors": errors}


@router.put("/question-bank/{question_id}")
async def update_question(question_id: str, body: dict, user=Depends(require_admin)):
    existing = await question_banks_collection.find_one({"_id": question_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Question not found")
    merged = {**existing, **body, "_id": question_id}
    try:
        question = clean_question(merged)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    await question_banks_collection.update_one({"_id": question_id}, {"$set": question})
    return {"question": serialize_doc(question)}


@router.get("/assessment-configs")
async def list_assessment_configs(user=Depends(require_admin)):
    configs = await assessment_configs_collection.find().sort("created_at", -1).to_list(50)
    return {"configs": [serialize_doc(c) for c in configs]}


@router.post("/assessment-configs")
async def create_assessment_config(body: dict, user=Depends(require_admin)):
    name = body.get('name', '').strip()
    assessment_type = body.get('assessment_type', 'comprehensive').strip()
    if not name:
        raise HTTPException(status_code=400, detail="Assessment name is required")
    existing = await assessment_configs_collection.find_one({"name": name})
    if existing:
        raise HTTPException(status_code=409, detail="An assessment with this name already exists")
    config_id = f"acfg_{uuid.uuid4().hex[:12]}"
    doc = {
        "_id": config_id,
        "name": name,
        "assessment_type": assessment_type,
        "question_count": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await assessment_configs_collection.insert_one(doc)
    return {"config": serialize_doc(doc)}


@router.put("/assessment-configs/{config_id}")
async def update_assessment_config(config_id: str, body: dict, user=Depends(require_admin)):
    existing = await assessment_configs_collection.find_one({"_id": config_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Assessment config not found")
    updates = {}
    if body.get('name'):
        updates['name'] = body['name'].strip()
    if body.get('assessment_type'):
        updates['assessment_type'] = body['assessment_type'].strip()
    updates['updated_at'] = datetime.now(timezone.utc).isoformat()
    await assessment_configs_collection.update_one({"_id": config_id}, {"$set": updates})
    updated = await assessment_configs_collection.find_one({"_id": config_id})
    return {"config": serialize_doc(updated)}


@router.delete("/assessment-configs/{config_id}")
async def delete_assessment_config(config_id: str, user=Depends(require_admin)):
    existing = await assessment_configs_collection.find_one({"_id": config_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Assessment config not found")
    await question_banks_collection.update_many(
        {"assessment_id": config_id},
        {"$unset": {"assessment_id": ""}}
    )
    await assessment_configs_collection.delete_one({"_id": config_id})
    return {"message": "Assessment config deleted"}


@router.get("/dashboard")
async def school_dashboard(user=Depends(require_admin)):
    query = scoped_school_filter(user)
    total_students = await students_collection.count_documents(query)
    total_schools = await schools_collection.count_documents({"_id": query['school_id']}) if query.get('school_id') else await schools_collection.count_documents({})
    assessment_query = query.copy()
    total_assessments = await assessments_collection.count_documents(assessment_query)
    completed_assessments = await assessments_collection.count_documents({**assessment_query, "status": "completed"})

    pipeline = [
        {"$match": {"status": "completed", **assessment_query}},
        {"$group": {"_id": None, "avg_score": {"$avg": "$overall_score"}}},
    ]
    avg_result = await assessments_collection.aggregate(pipeline).to_list(1)
    avg_score = avg_result[0]['avg_score'] if avg_result else 0

    grade_pipeline = [
        {"$match": query},
        {"$group": {"_id": "$class_level", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
    ]
    by_grade = await students_collection.aggregate(grade_pipeline).to_list(20)

    return {
        "total_schools": total_schools,
        "total_students": total_students,
        "total_assessments": total_assessments,
        "completed_assessments": completed_assessments,
        "pending_assessments": total_assessments - completed_assessments,
        "average_score": round(avg_score, 1) if avg_score else 0,
        "by_grade": by_grade,
    }
