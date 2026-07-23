import csv
import io
import re
import secrets
import string
import uuid
from datetime import datetime, timezone
from typing import Dict, List, Tuple

from database import credentials_collection, students_collection, users_collection
from services.auth_service import hash_password


ROMAN_MAP = {
    'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5, 'VI': 6,
    'VII': 7, 'VIII': 8, 'IX': 9, 'X': 10, 'XI': 11, 'XII': 12,
}


def roman_to_int(value: str) -> int:
    """Convert Roman numeral string (I-XII) to integer."""
    v = value.strip().upper()
    if v in ROMAN_MAP:
        return ROMAN_MAP[v]
    return int(v)


def parse_class_section(raw: str) -> Tuple[int, str]:
    """Parse a class value like 'V-A', 'XII-B', '8', 'VII C' into (class_level, section)."""
    raw = raw.strip()
    if not raw:
        raise ValueError("Empty class value")

    raw_upper = raw.upper()

    for roman in sorted(ROMAN_MAP.keys(), key=len, reverse=True):
        if raw_upper.startswith(roman):
            rest = raw_upper[len(roman):].strip().lstrip('-').strip()
            section = rest if rest else ''
            return ROMAN_MAP[roman], section

    match = re.match(r'^(\d+)\s*[-.]?\s*([A-Za-z]?)$', raw)
    if match:
        class_num = int(match.group(1))
        section = match.group(2).upper() if match.group(2) else ''
        return class_num, section

    try:
        return int(float(raw)), ''
    except (ValueError, TypeError):
        raise ValueError(f"Cannot parse class value: '{raw}'")


KNOWN_HEADER_KEYWORDS = {'s. no', 'student name', 'class', 'mobile number', 'roll', 'gender', 'name'}


def _is_header_row(values: list[str]) -> bool:
    """Heuristic: a row is a header if most of its non-empty cells look like column names."""
    non_empty = [v.strip().lower() for v in values if v.strip()]
    if not non_empty:
        return False
    hits = sum(1 for v in non_empty if any(kw in v for kw in KNOWN_HEADER_KEYWORDS))
    return hits >= 2


def parse_tabular_file(file_content: bytes, filename: str = "") -> List[Dict]:
    """Parse CSV or Excel files, auto-detecting header rows even with preamble lines."""
    if filename.lower().endswith((".xlsx", ".xls")):
        return _parse_excel(file_content)
    return _parse_csv_smart(file_content)


def _parse_excel(file_content: bytes) -> List[Dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = 0
    for idx, row in enumerate(rows):
        str_row = [str(cell or "").strip() for cell in row]
        if _is_header_row(str_row):
            header_idx = idx
            break

    headers = [str(cell or "").strip() for cell in rows[header_idx]]
    parsed = []
    for row in rows[header_idx + 1:]:
        vals = ["" if cell is None else str(cell).strip() for cell in row]
        if not any(vals):
            continue
        parsed.append({
            headers[i]: vals[i]
            for i in range(min(len(headers), len(vals)))
            if headers[i]
        })
    return parsed


def _parse_csv_smart(file_content: bytes) -> List[Dict]:
    """Parse CSV with automatic preamble/header detection."""
    text = file_content.decode('utf-8-sig')
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    if not all_rows:
        return []

    header_idx = 0
    for idx, row in enumerate(all_rows):
        if _is_header_row(row):
            header_idx = idx
            break

    headers = [h.strip() for h in all_rows[header_idx]]

    parsed = []
    for row in all_rows[header_idx + 1:]:
        if not row or not any(cell.strip() for cell in row):
            continue
        record = {}
        for i in range(min(len(headers), len(row))):
            if headers[i]:
                record[headers[i]] = row[i].strip()
        if record:
            parsed.append(record)
    return parsed


def _column_alias(normalized: dict, *candidates) -> str:
    for c in candidates:
        val = normalized.get(c, '')
        if val:
            return val
    return ''


def parse_csv(file_content: bytes) -> List[Dict]:
    return parse_tabular_file(file_content, "upload.csv")


def validate_csv_rows(rows: List[Dict]) -> Tuple[List[Dict], List[str]]:
    """Validate and normalize rows from any supported CSV format."""
    valid = []
    errors = []

    if not rows:
        return valid, ['No rows found in file']

    sample = {k.strip().lower().replace(' ', '_'): v for k, v in rows[0].items()}
    has_student_name = 'student_name' in sample or 'name' in sample
    has_class = 'class' in sample or 'class_level' in sample

    if has_student_name and has_class:
        return _validate_school_roster(rows)

    required_columns = {'name', 'class'}
    actual_cols = set(k.strip().lower().replace(' ', '_') for k in rows[0].keys())
    missing = required_columns - actual_cols
    if missing:
        errors.append(f"Missing required columns: {', '.join(missing)}")
        return valid, errors

    for i, row in enumerate(rows):
        normalized = {k.strip().lower().replace(' ', '_'): str(v or '').strip() for k, v in row.items()}
        name = normalized.get('name', '').strip()
        class_raw = normalized.get('class', '').strip()

        if not name:
            errors.append(f"Row {i + 2}: Name is empty")
            continue

        try:
            class_num = int(float(class_raw))
            section = normalized.get('section', '')
        except (ValueError, TypeError):
            errors.append(f"Row {i + 2}: Invalid class value '{class_raw}'")
            continue

        valid.append({
            'name': name,
            'class_level': class_num,
            'section': section,
            'roll_number': normalized.get('roll_number', normalized.get('roll_no', '')),
            'username': normalized.get('username', ''),
            'gender': normalized.get('gender', ''),
            'phone': normalized.get('mobile_number', normalized.get('mobile', '')),
        })

    return valid, errors


def _validate_school_roster(rows: List[Dict]) -> Tuple[List[Dict], List[str]]:
    """Validate school roster CSV (Student Name, Class, Mobile Number format)."""
    valid = []
    errors = []

    for i, row in enumerate(rows):
        normalized = {
            k.strip().lower().replace(' ', '_').replace('.', ''): str(v or '').strip()
            for k, v in row.items()
        }

        name = _column_alias(normalized, 'student_name', 'name')
        class_raw = _column_alias(normalized, 'class', 'class_level', 'grade')

        if not name:
            errors.append(f"Row {i + 2}: Student name is empty")
            continue
        if not class_raw:
            errors.append(f"Row {i + 2}: Class is empty for '{name}'")
            continue

        try:
            class_level, section = parse_class_section(class_raw)
            if class_level < 1 or class_level > 12:
                errors.append(f"Row {i + 2}: Class must be 1-12 (got {class_level} for '{name}')")
                continue
        except (ValueError, TypeError) as e:
            errors.append(f"Row {i + 2}: Invalid class '{class_raw}' for '{name}': {e}")
            continue

        s_no = _column_alias(normalized, 's_no', 'serial_no', 'roll_number', 'roll_no')
        phone = _column_alias(normalized, 'mobile_number', 'mobile', 'phone', 'contact')
        gender = _column_alias(normalized, 'gender', 'sex')

        valid.append({
            'name': name.strip(),
            'class_level': class_level,
            'section': section,
            'roll_number': s_no,
            'username': '',
            'gender': gender,
            'phone': phone,
        })

    return valid, errors


def generate_password(name: str, class_level: int) -> str:
    alphabet = string.ascii_letters + string.digits
    token = ''.join(secrets.choice(alphabet) for _ in range(8))
    return f"VL-{class_level}-{token}"


def generate_username(name: str, class_level: int, index: int) -> str:
    clean_name = name.lower().replace(' ', '').replace('.', '')
    clean_name = ''.join(c for c in clean_name if c.isalnum()) or 'student'
    return f"{clean_name}{class_level}_{index}"


async def create_student_accounts(valid_rows: List[Dict], school_id: str, admin_id: str) -> Dict:
    credentials = []
    now = datetime.now(timezone.utc).isoformat()

    for i, row in enumerate(valid_rows):
        username = row.get('username') or generate_username(row['name'], row['class_level'], i + 1)
        password = generate_password(row['name'], row['class_level'])
        user_id = f"usr_{uuid.uuid4().hex[:12]}"
        student_id = f"stu_{uuid.uuid4().hex[:12]}"

        existing_user = await users_collection.find_one({"username": username})
        if existing_user:
            username = f"{username}_{uuid.uuid4().hex[:4]}"

        user_doc = {
            "_id": user_id,
            "username": username,
            "password": hash_password(password),
            "name": row['name'],
            "role": "student",
            "school_id": school_id,
            "student_id": student_id,
            "first_login": True,
            "is_active": True,
            "created_at": now,
        }
        await users_collection.insert_one(user_doc)

        student_doc = {
            "_id": student_id,
            "user_id": user_id,
            "school_id": school_id,
            "name": row['name'],
            "class_level": row['class_level'],
            "section": row.get('section', ''),
            "roll_number": row.get('roll_number', ''),
            "gender": row.get('gender', ''),
            "username": username,
            "phone": row.get('phone', ''),
            "assigned_by": admin_id,
            "created_at": now,
        }
        await students_collection.insert_one(student_doc)

        credentials.append({
            "name": row['name'],
            "username": username,
            "password": password,
            "class_level": row['class_level'],
            "section": row.get('section', ''),
            "student_id": student_id,
            "user_id": user_id,
        })

    batch_id = f"batch_{uuid.uuid4().hex[:8]}"
    await credentials_collection.insert_one({
        "_id": batch_id,
        "school_id": school_id,
        "created_by": admin_id,
        "credentials": credentials,
        "created_at": now,
    })

    return {"batch_id": batch_id, "credentials": credentials}
